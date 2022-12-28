import os, time
from openpyxl import load_workbook
from openpyxl.utils import rows_from_range

def copia():
    wb = load_workbook(filename='pandas_to_excel_no_index_header.xlsx')
    sheet = wb["Sheet1"]
    sheet2 = wb["Novembro"]

    mr = sheet.max_row
    mc = sheet.max_column

    for i in range (1, mr + 1):
        for j in range (1, mc + 1):
            # reading cell value from source excel file
            c = sheet.cell(row = i, column = j)
    
            # writing the read value to destination excel file
            sheet2.cell(row = i, column = j).value = c.value

    wb.save('pandas_to_excel_no_index_header.xlsx')