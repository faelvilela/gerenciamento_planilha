import PySimpleGUI as sg
import os, time
from openpyxl import load_workbook
import datetime, time
from datetime import date
from dateutil.relativedelta import relativedelta, MO, SU, TU, WE, TH, FR, SA
import pandas as pd
import os, sys
from mes import *
from mover import copia

def main():

    layout = [  [sg.Text('Preenchimento de planilha')],
                [sg.Listbox(values=['segunda', 'terca', 'quarta', 'quinta', 'sexta', 'sabado'], select_mode='extended', key='fac', size=(52, 6))],
                [sg.Text('Dados:'), sg.InputText()],
                [sg.Button('Start'), sg.Button('Gerar Mes'), sg.Button('Abrir planilha'), sg.Button('Exit')]  ]

    window = sg.Window('Preenchimento de planilha', layout)
    
    while True:
        global v
        event,v=window.read()
        global strx
        strx=""
        #Setar strx como o valor que foi selecionado na listbox
        try:
            for val in v['fac']:
                strx=strx+ " "+ val+","
        except TypeError:
            pass
        #print(event, v)
        if event == sg.WIN_CLOSED or event == 'Exit':
            break
        if event == 'Start':
            #print('o dia é:', strx[1:len(strx)-1])
            #print(v['fac'])
            preenchimento()
            #window.close()
        if event == 'Gerar Mes':
            gerar_meses()
        if event == 'Abrir planilha':
            os.startfile('pandas_to_excel_no_index_header.xlsx')
        elif event == '-FUNCTION COMPLETED-':
            sg.popup('Funcao completada')
    window.close()


def preenchimento():

    today = date.today()
    #Setando 'today' como o ultimo domingo  
    sunday = today + relativedelta(weekday=SU)
    #Definindo os dias da semana atuais em formato de data 
    monday = sunday + relativedelta(weekday=MO(-1))
    tuesday = sunday + relativedelta(weekday=TU(-1))
    wednesday = sunday + relativedelta(weekday=WE(-1))
    thursday = sunday + relativedelta(weekday=TH(-1))
    friday = sunday + relativedelta(weekday=FR(-1))
    saturday = today + relativedelta(weekday=SA(-1))#Sabado Atual
    saturday_2 = sunday + relativedelta(weekday=SA(-1))#Proximo Sabado
    #Inicializando dados que devem ser preenchidos
    seg_dados = '0 0 0 0 0 0 0 0 0 0 0 0 0'
    ter_dados = '0 0 0 0 0 0 0 0 0 0 0 0 0'
    qua_dados = '0 0 0 0 0 0 0 0 0 0 0 0 0'
    qui_dados = '0 0 0 0 0 0 0 0 0 0 0 0 0'
    sex_dados = '0 0 0 0 0 0 0 0 0 0 0 0 0'
    sab_dados = '0 0 0 0 0 0 0 0 0 0 0 0 0'

    if strx[1:len(strx)-1] == 'segunda':
        #Definindo os dados preenchidos como vetor
        seg_dados = v[0]
        dia = 1
    elif strx[1:len(strx)-1] == 'terca':
        ter_dados = v[0]
        dia = 2
    elif strx[1:len(strx)-1] == 'quarta':
        qua_dados = v[0]
        dia = 3
    elif strx[1:len(strx)-1] == 'quinta':
        qui_dados = v[0]
        dia = 4
    elif strx[1:len(strx)-1] == 'sexta':
        sex_dados = v[0]
        dia = 5
    elif strx[1:len(strx)-1] == 'sabado':
        sab_dados = v[0]
        dia = 6
    
    #segunda 
    seg_slice = seg_dados.split() #splitando a string em partes, para depois pegar as partes nescessarias
    seg_primeiro, seg_segundo, seg_terceiro, seg_quarto, seg_quinto, seg_penultimo, seg_ultimo = seg_slice[0], seg_slice[1], seg_slice[2], seg_slice[3], seg_slice[4], seg_slice[11], seg_slice[12]
    #terca
    ter_slice = ter_dados.split()
    ter_primeiro, ter_segundo, ter_terceiro, ter_quarto, ter_quinto, ter_penultimo, ter_ultimo = ter_slice[0], ter_slice[1], ter_slice[2], ter_slice[3], ter_slice[4], ter_slice[11], ter_slice[12]
    #quarta
    qua_slice = qua_dados.split()
    qua_primeiro, qua_segundo, qua_terceiro, qua_quarto, qua_quinto, qua_penultimo, qua_ultimo = qua_slice[0], qua_slice[1], qua_slice[2], qua_slice[3], qua_slice[4], qua_slice[11], qua_slice[12]
    #quinta
    qui_slice = qui_dados.split()
    qui_primeiro, qui_segundo, qui_terceiro, qui_quarto, qui_quinto, qui_penultimo, qui_ultimo = qui_slice[0], qui_slice[1], qui_slice[2], qui_slice[3], qui_slice[4], qui_slice[11], qui_slice[12]
    #sexta
    sex_slice = sex_dados.split()
    sex_primeiro, sex_segundo, sex_terceiro, sex_quarto, sex_quinto, sex_penultimo, sex_ultimo = sex_slice[0], sex_slice[1], sex_slice[2], sex_slice[3], sex_slice[4], sex_slice[11], sex_slice[12]
    #sabado
    sab_slice = sab_dados.split()
    sab_primeiro, sab_segundo, sab_terceiro, sab_quarto, sab_quinto, sab_penultimo, sab_ultimo = sab_slice[0], sab_slice[1], sab_slice[2], sab_slice[3], sab_slice[4], sab_slice[11], sab_slice[12]


    def escreve_segunda():
        sheet['C7']= seg_primeiro
        sheet['D7']= seg_segundo
        sheet['E7']= seg_terceiro
        sheet['F7']= seg_quarto
        sheet['G7']= seg_quinto
        sheet['N7']= seg_penultimo
        sheet['O7']= seg_ultimo
        #Sempre que os dados de segunda forem preenchidos as datas na planilha são atualizadas para a nova semana
        sheet['B7'] = monday
        sheet['B8'] = tuesday
        sheet['B9'] = wednesday
        sheet['B10'] = thursday
        sheet['B11'] = friday
        sheet['B12'] = saturday_2

    def escreve_terca():
        sheet['B8'] = tuesday
        sheet['C8']= ter_primeiro
        sheet['D8']= ter_segundo
        sheet['E8']= ter_terceiro
        sheet['F8']= ter_quarto
        sheet['G8']= ter_quinto
        sheet['N8']= ter_penultimo
        sheet['O8']= ter_ultimo

    def escreve_quarta():
        sheet['B9'] = wednesday
        sheet['C9']= qua_primeiro
        sheet['D9']= qua_segundo
        sheet['E9']= qua_terceiro
        sheet['F9']= qua_quarto
        sheet['G9']= qua_quinto
        sheet['N9']= qua_penultimo
        sheet['O9']= qua_ultimo

    def escreve_quinta():
        sheet['B10'] = thursday
        sheet['C10']= qui_primeiro
        sheet['D10']= qui_segundo
        sheet['E10']= qui_terceiro
        sheet['F10']= qui_quarto
        sheet['G10']= qui_quinto
        sheet['N10']= qui_penultimo
        sheet['O10']= qui_ultimo

    def escreve_sexta():
        sheet['B11'] = friday
        sheet['C11']= sex_primeiro
        sheet['D11']= sex_segundo
        sheet['E11']= sex_terceiro
        sheet['F11']= sex_quarto
        sheet['G11']= sex_quinto
        sheet['N11']= sex_penultimo
        sheet['O11']= sex_ultimo

    def escreve_sabado():
        sheet['B12'] = saturday
        sheet['C12']= sab_primeiro
        sheet['D12']= sab_segundo
        sheet['E12']= sab_terceiro
        sheet['F12']= sab_quarto
        sheet['G12']= sab_quinto
        sheet['N12']= sab_penultimo
        sheet['O12']= sab_ultimo

    if dia == 1:
        wb = load_workbook(filename='C:/Users/rafaelvilela/Desktop/MEGAsync/Code/planilha/pandas_to_excel_no_index_header.xlsx')
        sheet = wb.active
        escreve_segunda()
        #Apagando conteudo da semana passada
        for a in sheet['C8':'O12']:
            for cell in a:
                cell.value = None
        wb.save('pandas_to_excel_no_index_header.xlsx')
    elif dia == 2:
        wb = load_workbook(filename='pandas_to_excel_no_index_header.xlsx')
        sheet = wb.active
        escreve_terca()
        wb.save('pandas_to_excel_no_index_header.xlsx')
    elif dia == 3:
        wb = load_workbook(filename='C:/Users/rafaelvilela/Desktop/MEGAsync/Code/planilha/pandas_to_excel_no_index_header.xlsx')
        sheet = wb.active
        escreve_quarta()
        wb.save('pandas_to_excel_no_index_header.xlsx')
    elif dia == 4:
        wb = load_workbook(filename='pandas_to_excel_no_index_header.xlsx')
        sheet = wb.active
        escreve_quinta()
        wb.save('pandas_to_excel_no_index_header.xlsx')
    elif dia == 5:
        wb = load_workbook(filename='pandas_to_excel_no_index_header.xlsx')
        sheet = wb.active
        escreve_sexta()
        wb.save('pandas_to_excel_no_index_header.xlsx')
    elif dia == 6:
        wb = load_workbook(filename='pandas_to_excel_no_index_header.xlsx')
        sheet = wb.active
        escreve_sabado()
        wb.save('pandas_to_excel_no_index_header.xlsx')
        os.startfile('months.xlsx')

    os.startfile('pandas_to_excel_no_index_header.xlsx')

if __name__ == "__main__":
    main()